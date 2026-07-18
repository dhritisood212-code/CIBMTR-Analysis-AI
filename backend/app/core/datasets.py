"""Handling of user-uploaded session datasets.

Compliance rules enforced here:
  * Uploaded data is written ONLY to the run's ephemeral `_session/` dir (gitignored, purged on
    finalize) - never to the persistent, committed artifact area.
  * Size and zip-bomb caps, since this accepts arbitrary uploads on a public endpoint.
  * The data dictionary is converted to plain TEXT for the Interpreter; the raw file is treated
    as inert data, never executed or interpreted as instructions.
"""
from __future__ import annotations

import zipfile
from pathlib import Path
from typing import BinaryIO

# Caps for a public endpoint that executes code. CIBMTR public datasets are small (a CSV + a
# dictionary), so these are generous but bounded.
MAX_UPLOAD_BYTES = 150 * 1024 * 1024      # 150 MB per file
MAX_UNZIP_BYTES = 500 * 1024 * 1024       # 500 MB total extracted
MAX_UNZIP_FILES = 2000
_CHUNK = 1024 * 1024


class UploadTooLarge(ValueError):
    pass


class BadArchive(ValueError):
    pass


def save_upload(fileobj: BinaryIO, dest: Path, max_bytes: int = MAX_UPLOAD_BYTES) -> Path:
    """Stream an upload to `dest`, enforcing a size cap. Returns dest."""
    dest.parent.mkdir(parents=True, exist_ok=True)
    total = 0
    with open(dest, "wb") as out:
        while True:
            chunk = fileobj.read(_CHUNK)
            if not chunk:
                break
            total += len(chunk)
            if total > max_bytes:
                out.close()
                dest.unlink(missing_ok=True)
                raise UploadTooLarge(
                    f"upload exceeds {max_bytes // (1024*1024)} MB limit")
            out.write(chunk)
    return dest


def extract_dataset(zip_path: Path, dest_dir: Path) -> list[str]:
    """Safely extract a dataset zip into dest_dir. Guards against path traversal and zip bombs.
    Returns the list of extracted relative paths. Non-zip files are left as-is (returns [])."""
    if not zipfile.is_zipfile(zip_path):
        return []  # e.g. the user uploaded a bare .csv rather than a zip
    dest_dir.mkdir(parents=True, exist_ok=True)
    extracted: list[str] = []
    total = 0
    with zipfile.ZipFile(zip_path) as z:
        infos = z.infolist()
        if len(infos) > MAX_UNZIP_FILES:
            raise BadArchive(f"archive has too many entries ({len(infos)})")
        for info in infos:
            if info.is_dir():
                continue
            # Resolve the target and ensure it stays inside dest_dir (no ../ traversal).
            target = (dest_dir / info.filename).resolve()
            if not str(target).startswith(str(dest_dir.resolve())):
                raise BadArchive(f"unsafe path in archive: {info.filename}")
            total += info.file_size
            if total > MAX_UNZIP_BYTES:
                raise BadArchive("archive expands beyond the extraction size limit")
            target.parent.mkdir(parents=True, exist_ok=True)
            with z.open(info) as src, open(target, "wb") as out:
                while True:
                    chunk = src.read(_CHUNK)
                    if not chunk:
                        break
                    out.write(chunk)
            extracted.append(str(target.relative_to(dest_dir)))
    return extracted


def dictionary_to_text(path: Path, max_chars: int = 60_000) -> str:
    """Convert a data dictionary to plain text the Interpreter can read, regardless of format
    (.xlsx/.xlsm, .rtf, .csv/.tsv/.txt/.md). Best-effort; never raises on format."""
    suffix = path.suffix.lower()
    try:
        if suffix in (".xlsx", ".xlsm", ".xltx"):
            text = _xlsx_to_text(path)
        elif suffix == ".rtf":
            text = _rtf_to_text(path)
        else:
            text = path.read_text(errors="replace")
    except Exception as exc:  # noqa: BLE001 - degrade gracefully, the run can still proceed
        text = f"[could not parse dictionary {path.name}: {exc}]"
    return text if len(text) <= max_chars else text[:max_chars] + "\n...[truncated]..."


def _xlsx_to_text(path: Path) -> str:
    from openpyxl import load_workbook  # lazy import

    wb = load_workbook(path, read_only=True, data_only=True)
    out: list[str] = []
    for ws in wb.worksheets:
        out.append(f"# sheet: {ws.title}")
        for row in ws.iter_rows(values_only=True):
            cells = ["" if c is None else str(c) for c in row]
            if any(cells):
                out.append("\t".join(cells))
    return "\n".join(out)


def _rtf_to_text(path: Path) -> str:
    raw = path.read_text(errors="replace")
    try:
        from striprtf.striprtf import rtf_to_text  # lazy import

        return rtf_to_text(raw)
    except Exception:  # noqa: BLE001 - fall back to a crude strip if striprtf is unavailable
        import re

        no_groups = re.sub(r"\\[a-zA-Z]+-?\d* ?", " ", raw)
        no_braces = re.sub(r"[{}]", "", no_groups)
        return re.sub(r"[ \t]+", " ", no_braces)
